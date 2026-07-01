"""Excel metadata reader for local author samples."""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

MetadataRow = dict[str, str | None]

EXPECTED_COLUMNS = {
    "url": "url",
    "added_date": "added_date",
    "added date": "added_date",
    "category": "category",
    "tags": "tags",
    "heading": "heading",
    "title": "title",
    "meta description": "meta_description",
    "meta_description": "meta_description",
    "keywords": "keywords",
    "content": "content",
    "filename": "filename",
    "file name": "filename",
}


def normalize_column_name(value: object) -> str:
    """Normalize an Excel header into the internal metadata key format."""

    normalized = str(value or "").strip().lower().replace("\n", " ")
    normalized = " ".join(normalized.split())
    return EXPECTED_COLUMNS.get(normalized, normalized.replace(" ", "_"))


def read_metadata_rows(path: Path) -> list[MetadataRow]:
    """Read the first worksheet in an Excel metadata file."""

    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        workbook.close()
        return []

    headers = [normalize_column_name(header) for header in rows[0]]
    metadata_rows: list[MetadataRow] = []
    for row in rows[1:]:
        row_data: MetadataRow = {}
        has_value = False
        for index, header in enumerate(headers):
            if not header:
                continue
            value = _stringify_cell(row[index] if index < len(row) else None)
            if value is not None:
                has_value = True
            row_data[header] = value
        if has_value:
            metadata_rows.append(row_data)

    workbook.close()
    return metadata_rows


def _stringify_cell(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None
